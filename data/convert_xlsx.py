import re
from collections import defaultdict
from pathlib import Path
from typing import Container, Mapping, Optional, Sequence

import openpyxl
from natsort import natsorted


def read_xlsx(file_path: Path, ignored_sheets: Optional[Container] = None) -> Mapping[str, Mapping[str, list[str]]]:
    """
    Read an xlsx file and return the data in a dictionary

    Args:
        file_path (Path): The path to the xlsx file
        ignored_sheets (Optional[Container], optional): A list of sheets to ignore. Defaults to None.

    Returns:
        Mapping[str, Mapping[str, list[str]]]: A dictionary with the data
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    # Loop over all sheets

    data = defaultdict(lambda: defaultdict(list))

    for sheet in wb.sheetnames:
        if ignored_sheets and sheet in ignored_sheets:
            continue
        ws = wb[sheet]
        header = None
        for row in ws.iter_rows():
            # Loop over all columns
            if not header:
                header = row
            else:
                for i, cell in enumerate(row):
                    data[sheet][header[i].value].append(cell.value)

    return data


def spinque_link_to_data(spinque_link: str, inventory_number: Optional[str] = None) -> tuple[str, int]:
    """
    Extract the inventory number and page number from a spinque link

    Args:
        spinque_link (str): The spinque link
        inventory_number (Optional[str], optional): The inventory number obtained using another method. Defaults to None.

    Raises:
        ValueError: If the inventory number does not match with the spinque link

    Returns:
        tuple[str, int]: The inventory number and page number
    """
    split_spinque_link = spinque_link.split("/")
    if inventory_number and split_spinque_link[-2] != inventory_number:
        raise ValueError(f"Inventory number {inventory_number} does not match with spinque link {spinque_link}")
    else:
        inventory_number = split_spinque_link[-2]

    page_number = int(split_spinque_link[-1])
    return inventory_number, page_number


def extract_start_of_documents(data: Mapping, link_columns: Container = ["Start of document"]) -> Mapping:
    """
    Extract the start of documents from the data

    Args:
        data (Mapping): The data from the xlsx file
        link_columns (Container, optional): The columns that contain the links. Defaults to ["Start of document"].

    Returns:
        Mapping: A dictionary with the start of documents
    """
    start_of_document_dict = defaultdict(lambda: defaultdict(int))
    for sheet, values in data.items():
        inventory_number = sheet
        for key, value in values.items():
            value = [v for v in value if v is not None]
            if key in link_columns:
                for spinque_link in value:
                    inventory_number, page_number = spinque_link_to_data(spinque_link, inventory_number)
                    start_of_document_dict[inventory_number][page_number] = 1

    return start_of_document_dict


def path_to_inventory_page_number(path: Path) -> tuple[str, int, bool]:
    """
    Convert a path to an inventory number and page number, and check if it is a deelopname, and if it is a deelopname, skip it

    Args:
        path (Path): The path to the file

    Raises:
        ValueError: If the path does not match the expected format
        ValueError: If the inventory number in the dir does not match with the inventory number in the file

    Returns:
        tuple[str, int, bool]: The inventory number, page number, and whether to skip the file
    """
    inventory_number_dir = path.parent.name
    if check := re.match(r"(.+)_(.+)_(\d+)(_deelopname\d+)?", path.stem):
        inventory_number_file = check.group(2)
        if inventory_number_dir != inventory_number_file:
            raise ValueError(
                f"Inventory number in dir {inventory_number_dir} does not match with inventory number in file {inventory_number_file}. Path: {path}"
            )
        page_number = int(check.group(3))
        skip = False
        if check.group(4):
            # This means it is a deelopname
            skip = True
        return inventory_number_file, page_number, skip
    else:
        raise ValueError(f"Path {path} does not match the expected format")


def link_with_paths(
    xlsx_file: Path,
    paths: Sequence[Path],
    ignored_sheets: Optional[Container] = None,
    link_columns: list[str] = ["Start of document", "URL nieuw document op volgorde"],
) -> list[list[list[Path]]]:
    """
    Link the paths with the xlsx file

    Args:
        xlsx_file (Path): The path to the xlsx file
        paths (Sequence[Path]): The paths to the files
        ignored_sheets (Optional[Container], optional): The sheets to ignore. Defaults to None.
        link_columns (list[str], optional): The columns that contain the links. Defaults to ["Start of document", "URL nieuw document op volgorde"].

    Returns:
        list[list[list[Path]]]: A list of documents
    """
    data: Mapping[str, Mapping[str, list[Path]]] = read_xlsx(xlsx_file, ignored_sheets=ignored_sheets)
    start_of_documents: Mapping[str, Mapping[int, int]] = extract_start_of_documents(data, link_columns=link_columns)

    all_documents = []
    inventory_documents = []
    current_document = []
    current_inventory_number = None
    # Just sort the paths to make sure they are in the right order
    for path in natsorted(paths):
        inventory_number, page_number, skip = path_to_inventory_page_number(path)

        # If the inventory number is not in the start of documents, skip the file
        if inventory_number not in start_of_documents:
            continue

        # For the first document, set the current inventory number
        if current_inventory_number is None:
            current_inventory_number = inventory_number
            current_document = [path]
            continue

        # If the inventory number is different, start a new document
        if current_inventory_number != inventory_number:
            inventory_documents.append(current_document)
            all_documents.append(inventory_documents)
            inventory_documents = []
            current_document = [path]
            current_inventory_number = inventory_number
            continue

        # Check if the page number is the start of a document
        if skip:
            is_start = False
        else:
            try:
                is_start = start_of_documents[inventory_number][page_number]
            except KeyError:
                is_start = False

        # If it is the start of a document, start a new document
        if is_start:
            inventory_documents.append(current_document)
            current_document = [path]
        else:
            current_document.append(path)

    # Add the last document
    if current_document:
        inventory_documents.append(current_document)
        all_documents.append(inventory_documents)

    return all_documents


if __name__ == "__main__":
    paths = list(Path("~/Downloads/ushmm_test/").expanduser().rglob("*.jpg"))
    paths = natsorted(paths)
    print(len(paths))
    documents = link_with_paths(Path("/home/stefan/Downloads/Documentsegmentatie_GT.xlsx"), paths)
    print(documents)
    for document in documents:
        print(len(document))

    print(sum(len(document) for document in documents))
